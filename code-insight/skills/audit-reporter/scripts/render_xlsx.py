"""Render code-audit findings into an Excel workbook.

The workbook is the only report export for the code-insight audit pipeline.
It supports two lifecycle points:

- ``draft``: after one or more reviewer modules have written ``findings/*.json``.
- ``final``: after verifier results have refreshed ``findings/*.json`` and
  written ``confirmed.json``.

``auto`` mode renders ``final`` when ``confirmed.json`` exists, otherwise
renders ``draft``. The output path defaults to
``.garage/code-audit/runs/<run_id>/reports/report.xlsx``.
"""

from __future__ import annotations

import argparse
import dataclasses
import datetime as dt
import json
import sys
from collections import Counter
from pathlib import Path
from typing import Any

PACK_VERSION = "0.4.0"

VALID_SEVERITIES = ("critical", "high", "medium", "low", "info")
VALID_CONFIDENCES = ("high", "medium", "low")
BASE_11_CATEGORIES = (
    "correctness",
    "error-handling",
    "concurrency",
    "resource-leak",
    "security",
    "api-misuse",
    "typing",
    "performance",
    "dead-code",
    "contract-violation",
    "i18n-or-encoding",
)
VALID_CATEGORIES = BASE_11_CATEGORIES
VALID_VERIFIER_STATUSES = (
    "confirmed",
    "rejected",
    "upgrade",
    "downgrade",
    "needs_more_evidence",
)
CONFIRMED_STATUSES = ("confirmed", "upgrade", "downgrade")
REPORT_MODES = ("auto", "draft", "final")

SEVERITY_ZH = {
    "critical": "严重",
    "high": "高",
    "medium": "中",
    "low": "低",
    "info": "提示",
}
CONFIDENCE_ZH = {
    "high": "高",
    "medium": "中",
    "low": "低",
}
STATUS_ZH = {
    "draft": "待复核",
    "confirmed": "已确认",
    "rejected": "非问题",
    "upgrade": "已确认-严重级别上调",
    "downgrade": "已确认-严重级别下调",
    "needs_more_evidence": "证据不足",
}

HEADER_FILL_COLOR = "FF34495E"
SECTION_FILL_COLOR = "FFEEF2F7"
SEVERITY_FILL_COLORS = {
    "critical": "FFC0392B",
    "high": "FFE67E22",
    "medium": "FFF39C12",
    "low": "FF27AE60",
    "info": "FF7F8C8D",
}
STATUS_FILL_COLORS = {
    "draft": "FFBDC3C7",
    "confirmed": "FF27AE60",
    "rejected": "FF7F8C8D",
    "upgrade": "FFC0392B",
    "downgrade": "FFF39C12",
    "needs_more_evidence": "FF8E44AD",
}
CODE_SNIPPET_MAX_CHARS = 500

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    _OPENPYXL_AVAILABLE = True
    _OPENPYXL_IMPORT_ERROR: str | None = None
except ImportError as exc:  # pragma: no cover - env-dependent
    Workbook = None  # type: ignore[assignment,misc]
    Alignment = None  # type: ignore[assignment,misc]
    Font = None  # type: ignore[assignment,misc]
    PatternFill = None  # type: ignore[assignment,misc]
    get_column_letter = None  # type: ignore[assignment]
    _OPENPYXL_AVAILABLE = False
    _OPENPYXL_IMPORT_ERROR = str(exc)


class ReportError(ValueError):
    """Raised when input JSON is missing required fields or has invalid values."""


@dataclasses.dataclass(frozen=True)
class XlsxResult:
    output_path: Path
    mode: str
    finding_count: int
    confirmed_count: int
    by_severity: dict[str, int]
    by_module: dict[str, int]
    rejected_count: int
    needs_more_evidence_count: int
    skipped: bool = False
    skipped_reason: str | None = None


FINDING_COLUMNS = [
    ("发现ID", 22),
    ("模块", 18),
    ("文件", 50),
    ("起始行", 10),
    ("结束行", 10),
    ("位置", 36),
    ("标题", 50),
    ("审查类别", 22),
    ("审查类别说明", 44),
    ("严重级别", 12),
    ("严重级别说明", 16),
    ("置信度", 10),
    ("置信度说明", 14),
    ("复核状态", 24),
    ("复核状态说明", 26),
    ("问题描述", 60),
    ("证据代码", 60),
    ("审查推理", 60),
    ("触发条件", 44),
    ("预期与实际", 44),
    ("建议修复", 52),
    ("一审Agent", 28),
    ("一审时间", 22),
    ("复核理由", 56),
    ("复核核验证据", 56),
    ("调整前严重级别", 16),
]


NON_ISSUE_COLUMNS = [
    ("发现ID", 22),
    ("模块", 18),
    ("文件", 50),
    ("位置", 36),
    ("标题", 50),
    ("审查类别", 22),
    ("严重级别", 12),
    ("复核状态", 24),
    ("复核理由", 60),
    ("复核核验证据", 60),
    ("原问题描述", 60),
]

ISSUE_SUMMARY_COLUMNS = [
    ("发现ID", 22),
    ("模块", 18),
    ("位置", 36),
    ("审查类别", 22),
    ("严重级别", 12),
    ("复核状态", 24),
    ("标题", 50),
    ("问题说明", 70),
    ("建议修复", 60),
]


def render_workbook(
    *,
    workspace_root: Path | None = None,
    run_id: str | None = None,
    confirmed_path: Path | None = None,
    plan_path: Path | None = None,
    findings_dir: Path | None = None,
    output_path: Path | None = None,
    mode: str = "auto",
    strict: bool = True,
) -> XlsxResult:
    """Render an Excel workbook.

    ``draft`` mode only requires ``findings/*.json`` and accepts ``verifier: {}``.
    ``final`` mode requires verifier fields on every finding and uses
    ``confirmed.json`` to cross-check confirmed / upgraded / downgraded rows.
    """
    if mode not in REPORT_MODES:
        raise ReportError(f"mode must be one of {REPORT_MODES}, got {mode!r}")

    confirmed_path, plan_path, findings_dir, output_path = _resolve_paths(
        workspace_root=workspace_root,
        run_id=run_id,
        confirmed_path=confirmed_path,
        plan_path=plan_path,
        findings_dir=findings_dir,
        output_path=output_path,
    )
    actual_mode = _resolve_mode(mode, confirmed_path)

    plan = _load_json(plan_path) if plan_path.is_file() else {}
    allowed_categories = derive_allowed_categories(plan)
    category_descriptions = derive_category_descriptions(plan)

    if not findings_dir.is_dir():
        raise ReportError(f"Missing required directory: {findings_dir}")
    all_findings = _load_all_findings(findings_dir)

    _validate_findings(
        all_findings,
        allowed_categories=allowed_categories,
        require_verifier=(actual_mode == "final"),
    )

    confirmed = _load_confirmed(confirmed_path) if actual_mode == "final" else []
    if confirmed:
        _validate_findings(
            confirmed,
            allowed_categories=allowed_categories,
            require_verifier=True,
        )
        _validate_confirmed_consistency(all_findings, confirmed)

    confirmed_records = [
        f for f in all_findings if _verifier_status(f) in CONFIRMED_STATUSES
    ]
    non_issues = [f for f in all_findings if _verifier_status(f) == "rejected"]
    needs_more_evidence = [
        f for f in all_findings if _verifier_status(f) == "needs_more_evidence"
    ]

    by_severity = Counter(f["severity"] for f in all_findings)
    by_module = Counter(f["module"] for f in all_findings)

    if not _OPENPYXL_AVAILABLE:
        reason = f"openpyxl not installed ({_OPENPYXL_IMPORT_ERROR})"
        if strict:
            raise ReportError(reason)
        return XlsxResult(
            output_path=output_path,
            mode=actual_mode,
            finding_count=len(all_findings),
            confirmed_count=len(confirmed_records),
            by_severity=dict(by_severity),
            by_module=dict(by_module),
            rejected_count=len(non_issues),
            needs_more_evidence_count=len(needs_more_evidence),
            skipped=True,
            skipped_reason=reason,
        )

    wb = Workbook()
    ws_findings = wb.active
    ws_findings.title = "审查结果"
    _build_findings_sheet(
        ws_findings,
        all_findings,
        category_descriptions=category_descriptions,
    )

    ws_issue_summary = wb.create_sheet("问题总结")
    _build_issue_summary_sheet(
        ws_issue_summary,
        plan,
        all_findings,
        actual_mode,
        category_descriptions=category_descriptions,
    )

    ws_summary = wb.create_sheet("汇总")
    _build_summary_sheet(ws_summary, plan, all_findings, actual_mode)

    ws_runmeta = wb.create_sheet("运行信息")
    _build_runmeta_sheet(
        ws_runmeta,
        plan,
        all_findings,
        confirmed_records,
        non_issues,
        needs_more_evidence,
        actual_mode,
    )

    ws_non_issues = wb.create_sheet("非问题记录")
    _build_non_issue_sheet(
        ws_non_issues,
        non_issues,
        category_descriptions=category_descriptions,
    )

    ws_needs_more = wb.create_sheet("待补证据")
    _build_non_issue_sheet(
        ws_needs_more,
        needs_more_evidence,
        category_descriptions=category_descriptions,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return XlsxResult(
        output_path=output_path,
        mode=actual_mode,
        finding_count=len(all_findings),
        confirmed_count=len(confirmed_records),
        by_severity=dict(by_severity),
        by_module=dict(by_module),
        rejected_count=len(non_issues),
        needs_more_evidence_count=len(needs_more_evidence),
        skipped=False,
    )


def _resolve_paths(
    *,
    workspace_root: Path | None,
    run_id: str | None,
    confirmed_path: Path | None,
    plan_path: Path | None,
    findings_dir: Path | None,
    output_path: Path | None,
) -> tuple[Path, Path, Path, Path]:
    if output_path and (confirmed_path or findings_dir):
        base_dir = (
            confirmed_path.parent
            if confirmed_path is not None
            else findings_dir.parent
        )
        return (
            confirmed_path or (base_dir / "confirmed.json"),
            plan_path or (base_dir / "plan.json"),
            findings_dir or (base_dir / "findings"),
            output_path,
        )

    if workspace_root is None or run_id is None:
        raise ReportError(
            "Either (workspace_root + run_id) or (output_path + confirmed_path/findings_dir) must be provided"
        )

    run_dir = workspace_root / ".garage" / "code-audit" / "runs" / run_id
    return (
        confirmed_path or (run_dir / "confirmed.json"),
        plan_path or (run_dir / "plan.json"),
        findings_dir or (run_dir / "findings"),
        output_path or (run_dir / "reports" / "report.xlsx"),
    )


def _resolve_mode(mode: str, confirmed_path: Path) -> str:
    if mode == "auto":
        return "final" if confirmed_path.is_file() else "draft"
    return mode


def _load_json(path: Path) -> Any:
    if not path.is_file():
        raise ReportError(f"Missing required file: {path}")
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise ReportError(f"Invalid JSON in {path}: {exc}") from exc


def _load_confirmed(path: Path) -> list[dict[str, Any]]:
    confirmed = _load_json(path)
    if not isinstance(confirmed, list):
        raise ReportError(f"{path} must contain a JSON array of findings")
    return confirmed


def _load_all_findings(findings_dir: Path) -> list[dict[str, Any]]:
    if not findings_dir.is_dir():
        return []
    out: list[dict[str, Any]] = []
    for p in sorted(findings_dir.glob("*.json")):
        arr = _load_json(p)
        if not isinstance(arr, list):
            raise ReportError(f"{p} must contain a JSON array of findings")
        for item in arr:
            if not isinstance(item, dict):
                raise ReportError(f"{p} contains a non-object finding: {type(item).__name__}")
            out.append(item)
    return out


def derive_allowed_categories(plan: dict[str, Any]) -> tuple[str, ...]:
    rc = plan.get("review_checklist") if isinstance(plan, dict) else None
    if isinstance(rc, dict):
        cats = rc.get("categories")
        if isinstance(cats, list) and cats:
            ids: list[str] = []
            for entry in cats:
                if isinstance(entry, dict) and isinstance(entry.get("id"), str):
                    cid = entry["id"].strip()
                    if cid and cid not in ids:
                        ids.append(cid)
            if ids:
                return tuple(ids)
    return BASE_11_CATEGORIES


def derive_category_descriptions(plan: dict[str, Any]) -> dict[str, str]:
    rc = plan.get("review_checklist") if isinstance(plan, dict) else None
    if not isinstance(rc, dict):
        return {}
    cats = rc.get("categories")
    if not isinstance(cats, list):
        return {}
    out: dict[str, str] = {}
    for entry in cats:
        if not isinstance(entry, dict):
            continue
        cid = entry.get("id")
        if not isinstance(cid, str) or not cid.strip():
            continue
        desc = entry.get("description")
        out[cid.strip()] = str(desc or "")
    return out


REQUIRED_FINDING_FIELDS = (
    "id",
    "module",
    "file",
    "line_start",
    "line_end",
    "file_sha256",
    "title",
    "category",
    "severity",
    "confidence",
    "description",
    "evidence",
    "suggested_fix",
    "reviewer",
    "verifier",
)
REQUIRED_EVIDENCE_FIELDS = (
    "code_snippet",
    "reasoning",
    "trigger_conditions",
    "expected_vs_actual",
)
REQUIRED_REVIEWER_FIELDS = ("agent", "ts")
REQUIRED_VERIFIER_FIELDS = ("status", "reason", "evidence_check", "agent", "ts")


def _validate_findings(
    findings: list[Any],
    *,
    allowed_categories: tuple[str, ...] | None = None,
    require_verifier: bool,
) -> None:
    cats = tuple(allowed_categories) if allowed_categories else BASE_11_CATEGORIES
    for i, f in enumerate(findings):
        if not isinstance(f, dict):
            raise ReportError(f"finding #{i} must be an object, got {type(f).__name__}")
        for k in REQUIRED_FINDING_FIELDS:
            if k not in f:
                raise ReportError(f"finding #{i} ({f.get('id', '<no-id>')}) missing field '{k}'")
        if f["category"] not in cats:
            raise ReportError(
                f"finding {f['id']} has invalid category {f['category']!r}; must be one of {cats}"
            )
        if f["severity"] not in VALID_SEVERITIES:
            raise ReportError(
                f"finding {f['id']} has invalid severity {f['severity']!r}; must be one of {VALID_SEVERITIES}"
            )
        if f["confidence"] not in VALID_CONFIDENCES:
            raise ReportError(
                f"finding {f['id']} has invalid confidence {f['confidence']!r}; must be one of {VALID_CONFIDENCES}"
            )
        if not isinstance(f["line_start"], int) or not isinstance(f["line_end"], int):
            raise ReportError(f"finding {f['id']} line_start/line_end must be integers")
        if f["line_start"] > f["line_end"]:
            raise ReportError(
                f"finding {f['id']} line_start={f['line_start']} > line_end={f['line_end']}"
            )
        ev = f["evidence"]
        if not isinstance(ev, dict):
            raise ReportError(f"finding {f['id']} evidence must be an object")
        for k in REQUIRED_EVIDENCE_FIELDS:
            if k not in ev:
                raise ReportError(f"finding {f['id']} evidence missing field '{k}'")
        _validate_chinese_finding_text(f)
        rv = f["reviewer"]
        if not isinstance(rv, dict):
            raise ReportError(f"finding {f['id']} reviewer must be an object")
        for k in REQUIRED_REVIEWER_FIELDS:
            if k not in rv:
                raise ReportError(f"finding {f['id']} reviewer missing field '{k}'")
        _validate_verifier(f, require_verifier=require_verifier)


def _validate_verifier(finding: dict[str, Any], *, require_verifier: bool) -> None:
    vr = finding["verifier"]
    if not isinstance(vr, dict):
        raise ReportError(f"finding {finding['id']} verifier must be an object")
    if not vr and not require_verifier:
        return
    for k in REQUIRED_VERIFIER_FIELDS:
        if k not in vr:
            raise ReportError(f"finding {finding['id']} verifier missing field '{k}'")
    if vr["status"] not in VALID_VERIFIER_STATUSES:
        raise ReportError(
            f"finding {finding['id']} verifier.status {vr['status']!r} invalid; must be one of {VALID_VERIFIER_STATUSES}"
        )
    _require_chinese_text(vr["reason"], f"finding {finding['id']} verifier.reason")
    _require_chinese_text(
        vr["evidence_check"],
        f"finding {finding['id']} verifier.evidence_check",
    )
    severity_after = vr.get("severity_after")
    if severity_after is not None and severity_after not in VALID_SEVERITIES:
        raise ReportError(
            f"finding {finding['id']} verifier.severity_after {severity_after!r} invalid; must be one of {VALID_SEVERITIES}"
        )


def _validate_chinese_finding_text(finding: dict[str, Any]) -> None:
    evidence = finding["evidence"]
    for field in ("title", "description", "suggested_fix"):
        _require_chinese_text(finding[field], f"finding {finding['id']} {field}")
    for field in ("reasoning", "trigger_conditions", "expected_vs_actual"):
        _require_chinese_text(
            evidence[field],
            f"finding {finding['id']} evidence.{field}",
        )


def _require_chinese_text(value: Any, field_name: str) -> None:
    text = str(value or "").strip()
    if not text:
        raise ReportError(f"{field_name} must be non-empty Chinese text")
    if not any("\u4e00" <= ch <= "\u9fff" for ch in text):
        raise ReportError(
            f"{field_name} must contain Chinese explanatory text; "
            "English identifiers, paths, and API names may be included as supporting terms"
        )


def _validate_confirmed_consistency(
    all_findings: list[dict[str, Any]],
    confirmed: list[dict[str, Any]],
) -> None:
    expected = {f["id"] for f in all_findings if _verifier_status(f) in CONFIRMED_STATUSES}
    actual = {f["id"] for f in confirmed}
    if expected != actual:
        missing = sorted(expected - actual)
        extra = sorted(actual - expected)
        raise ReportError(
            "confirmed.json does not match findings verifier statuses; "
            f"missing={missing}, extra={extra}"
        )


def _truncate(text: Any, max_chars: int) -> str:
    s = str(text or "")
    if len(s) <= max_chars:
        return s
    return s[: max_chars - 1] + "…"


def _location(finding: dict[str, Any]) -> str:
    return f"{finding['file']}:{finding['line_start']}-{finding['line_end']}"


def _verifier_status(finding: dict[str, Any]) -> str:
    vr = finding.get("verifier") or {}
    if not isinstance(vr, dict) or not vr.get("status"):
        return "draft"
    return str(vr["status"])


def _verifier_reason(finding: dict[str, Any]) -> str:
    vr = finding.get("verifier") or {}
    return str(vr.get("reason") or "") if isinstance(vr, dict) else ""


def _verifier_evidence_check(finding: dict[str, Any]) -> str:
    vr = finding.get("verifier") or {}
    return str(vr.get("evidence_check") or "") if isinstance(vr, dict) else ""


def _sorted(findings: list[dict[str, Any]]) -> list[dict[str, Any]]:
    severity_order = {sev: i for i, sev in enumerate(VALID_SEVERITIES)}
    return sorted(
        findings,
        key=lambda f: (
            severity_order.get(f["severity"], 99),
            f["module"],
            f["file"],
            int(f["line_start"]),
            f["id"],
        ),
    )


def _build_findings_sheet(
    ws: Any,
    findings: list[dict[str, Any]],
    *,
    category_descriptions: dict[str, str],
) -> None:
    _write_header(ws, FINDING_COLUMNS)
    ws.freeze_panes = "A2"

    severity_fills = {
        sev: PatternFill("solid", fgColor=color)
        for sev, color in SEVERITY_FILL_COLORS.items()
    }
    status_fills = {
        status: PatternFill("solid", fgColor=color)
        for status, color in STATUS_FILL_COLORS.items()
    }
    light_font = Font(color="FFFFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    for row_idx, finding in enumerate(_sorted(findings), start=2):
        evidence = finding["evidence"]
        reviewer = finding["reviewer"]
        verifier = finding.get("verifier") or {}
        status = _verifier_status(finding)
        values = [
            finding["id"],
            finding["module"],
            finding["file"],
            finding["line_start"],
            finding["line_end"],
            _location(finding),
            finding["title"],
            finding["category"],
            category_descriptions.get(finding["category"], ""),
            finding["severity"],
            SEVERITY_ZH.get(finding["severity"], finding["severity"]),
            finding["confidence"],
            CONFIDENCE_ZH.get(finding["confidence"], finding["confidence"]),
            status,
            STATUS_ZH.get(status, status),
            finding["description"],
            _truncate(evidence["code_snippet"], CODE_SNIPPET_MAX_CHARS),
            evidence["reasoning"],
            evidence["trigger_conditions"],
            evidence["expected_vs_actual"],
            finding["suggested_fix"],
            reviewer["agent"],
            reviewer["ts"],
            _verifier_reason(finding),
            _verifier_evidence_check(finding),
            finding.get("severity_before") or verifier.get("severity_before") or "",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap

        severity_cell = ws.cell(row=row_idx, column=_column_index(FINDING_COLUMNS, "严重级别"))
        severity_cell.fill = severity_fills.get(finding["severity"], severity_fills["info"])
        severity_cell.font = light_font

        status_cell = ws.cell(row=row_idx, column=_column_index(FINDING_COLUMNS, "复核状态"))
        status_cell.fill = status_fills.get(status, status_fills["draft"])
        status_cell.font = light_font

    ws.auto_filter.ref = ws.dimensions


def _build_summary_sheet(
    ws: Any,
    plan: dict[str, Any],
    findings: list[dict[str, Any]],
    mode: str,
) -> None:
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    section_fill = PatternFill("solid", fgColor=SECTION_FILL_COLOR)
    bold = Font(bold=True)

    row = 1
    row = _write_section_title(ws, row, f"报告模式: {STATUS_ZH['draft'] if mode == 'draft' else '复核后最终结果'}")
    row = _write_counter_block(ws, row, "按复核状态统计", _status_counter(findings), STATUS_ZH)
    row += 1
    row = _write_counter_block(ws, row, "按严重级别统计", Counter(f["severity"] for f in findings), SEVERITY_ZH)
    row += 1
    row = _write_counter_block(ws, row, "按审查类别统计", Counter(f["category"] for f in findings), {})
    row += 1

    ws.cell(row=row, column=1, value="按模块和严重级别统计").font = bold
    ws.cell(row=row, column=1).fill = section_fill
    row += 1

    modules_in_plan = [m.get("name", "<unknown>") for m in plan.get("modules", [])]
    modules_seen = list(dict.fromkeys(modules_in_plan + sorted({f["module"] for f in findings})))
    if modules_seen:
        ws.cell(row=row, column=1, value="严重级别").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        for c_idx, module in enumerate(modules_seen, start=2):
            cell = ws.cell(row=row, column=c_idx, value=module)
            cell.font = header_font
            cell.fill = header_fill
        total_col = len(modules_seen) + 2
        cell = ws.cell(row=row, column=total_col, value="合计")
        cell.font = header_font
        cell.fill = header_fill
        row += 1

        by_module_severity: dict[str, Counter[str]] = {m: Counter() for m in modules_seen}
        for finding in findings:
            by_module_severity.setdefault(finding["module"], Counter())[finding["severity"]] += 1

        for severity in VALID_SEVERITIES:
            ws.cell(row=row, column=1, value=f"{severity} / {SEVERITY_ZH[severity]}").font = bold
            row_total = 0
            for c_idx, module in enumerate(modules_seen, start=2):
                count = by_module_severity.get(module, Counter()).get(severity, 0)
                ws.cell(row=row, column=c_idx, value=count)
                row_total += count
            ws.cell(row=row, column=total_col, value=row_total).font = bold
            row += 1

        ws.cell(row=row, column=1, value="合计").font = bold
        grand_total = 0
        for c_idx, module in enumerate(modules_seen, start=2):
            total = sum(by_module_severity.get(module, Counter()).values())
            ws.cell(row=row, column=c_idx, value=total).font = bold
            grand_total += total
        ws.cell(row=row, column=total_col, value=grand_total).font = bold

    for c_idx in range(1, max(8, ws.max_column) + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 18


def _build_issue_summary_sheet(
    ws: Any,
    plan: dict[str, Any],
    findings: list[dict[str, Any]],
    mode: str,
    *,
    category_descriptions: dict[str, str],
) -> None:
    issue_findings = _issue_findings_for_summary(findings, mode)
    non_issue_count = sum(1 for finding in findings if _verifier_status(finding) == "rejected")
    needs_more_evidence_count = sum(
        1 for finding in findings if _verifier_status(finding) == "needs_more_evidence"
    )
    wrap = Alignment(wrap_text=True, vertical="top")

    row = 1
    row = _write_section_title(ws, row, "问题总结")
    overview_rows = [
        ("报告模式", "一审草稿" if mode == "draft" else "复核后最终结果"),
        ("审查目标", str(plan.get("target") or "")),
        ("纳入本页总结的问题数", len(issue_findings)),
        ("全部 finding 数", len(findings)),
        ("复核认为非问题数", non_issue_count),
        ("仍需补证据数", needs_more_evidence_count),
    ]
    for key, value in overview_rows:
        ws.cell(row=row, column=1, value=key).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value).alignment = wrap
        row += 1

    row += 1
    row = _write_counter_block(
        ws,
        row,
        "本页问题按严重级别统计",
        Counter(finding["severity"] for finding in issue_findings),
        SEVERITY_ZH,
    )
    row += 1
    row = _write_counter_block(
        ws,
        row,
        "本页问题按模块统计",
        Counter(finding["module"] for finding in issue_findings),
        {},
    )
    row += 1
    row = _write_counter_block(
        ws,
        row,
        "本页问题按审查类别统计",
        Counter(finding["category"] for finding in issue_findings),
        category_descriptions,
    )
    row += 1

    row = _write_section_title(ws, row, "重点问题列表")
    _write_table_header(ws, row, ISSUE_SUMMARY_COLUMNS)
    row += 1
    for finding in _sorted(issue_findings):
        values = [
            finding["id"],
            finding["module"],
            _location(finding),
            finding["category"],
            finding["severity"],
            STATUS_ZH.get(_verifier_status(finding), _verifier_status(finding)),
            finding["title"],
            finding["description"],
            finding["suggested_fix"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.alignment = wrap
        severity_cell = ws.cell(row=row, column=_column_index(ISSUE_SUMMARY_COLUMNS, "严重级别"))
        severity_cell.fill = PatternFill(
            "solid",
            fgColor=SEVERITY_FILL_COLORS.get(finding["severity"], SEVERITY_FILL_COLORS["info"]),
        )
        severity_cell.font = Font(color="FFFFFFFF", bold=True)
        row += 1
    if not issue_findings:
        ws.cell(row=row, column=1, value="无已确认问题" if mode == "final" else "无 finding 草稿")


def _issue_findings_for_summary(
    findings: list[dict[str, Any]],
    mode: str,
) -> list[dict[str, Any]]:
    if mode == "draft":
        return list(findings)
    return [
        finding
        for finding in findings
        if _verifier_status(finding) in CONFIRMED_STATUSES
    ]


def _write_section_title(ws: Any, row: int, title: str) -> int:
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor=SECTION_FILL_COLOR)
    return row + 1


def _write_counter_block(
    ws: Any,
    row: int,
    title: str,
    counts: Counter[str],
    labels: dict[str, str],
) -> int:
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    row = _write_section_title(ws, row, title)
    headers = ("字段值", "中文说明", "数量")
    for c_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    row += 1
    for key, count in sorted(counts.items()):
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=labels.get(key, ""))
        ws.cell(row=row, column=3, value=count)
        row += 1
    if not counts:
        ws.cell(row=row, column=1, value="无")
        ws.cell(row=row, column=3, value=0)
        row += 1
    return row


def _build_runmeta_sheet(
    ws: Any,
    plan: dict[str, Any],
    findings: list[dict[str, Any]],
    confirmed_records: list[dict[str, Any]],
    non_issues: list[dict[str, Any]],
    needs_more_evidence: list[dict[str, Any]],
    mode: str,
) -> None:
    _write_header(ws, [("字段", 30), ("值", 70)])

    profile = plan.get("profile") if isinstance(plan.get("profile"), dict) else {}
    checklist = (
        plan.get("review_checklist")
        if isinstance(plan.get("review_checklist"), dict)
        else {}
    )

    def _list_str(value: Any) -> str:
        return ", ".join(str(v) for v in value) if isinstance(value, list) else ""

    rows: list[tuple[str, Any]] = [
        ("报告模式", "一审草稿" if mode == "draft" else "复核后最终结果"),
        ("run_id", str(plan.get("run_id") or "")),
        ("审查目标", str(plan.get("target") or "")),
        ("生成时间", dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds")),
        ("pack_version", PACK_VERSION),
        ("项目语言", _list_str(profile.get("languages"))),
        ("项目架构", _list_str(profile.get("architectures"))),
        ("框架", _list_str(profile.get("frameworks"))),
        ("风险关注点", _list_str(profile.get("risk_focus"))),
        ("项目画像是否人工确认", str(profile.get("user_confirmed", ""))),
        ("审查清单 preset", str(checklist.get("preset") or "")),
        ("审查类别数量", len(checklist.get("categories", [])) if isinstance(checklist.get("categories"), list) else 0),
        ("审查清单是否人工确认", str(checklist.get("user_confirmed", ""))),
        ("一审 finding 总数", len(findings)),
        ("复核确认问题数", len(confirmed_records)),
        ("复核认为非问题数", len(non_issues)),
        ("仍需补证据数", len(needs_more_evidence)),
        ("模块数量", len(plan.get("modules", [])) if isinstance(plan.get("modules"), list) else 0),
    ]

    for severity in VALID_SEVERITIES:
        rows.append((f"按严重级别.{severity}.{SEVERITY_ZH[severity]}", Counter(f["severity"] for f in findings).get(severity, 0)))
    for status, count in _status_counter(findings).items():
        rows.append((f"按复核状态.{status}.{STATUS_ZH.get(status, status)}", count))

    wrap = Alignment(wrap_text=True, vertical="top")
    for row_idx, (key, value) in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=value).alignment = wrap


def _build_non_issue_sheet(
    ws: Any,
    findings: list[dict[str, Any]],
    *,
    category_descriptions: dict[str, str],
) -> None:
    _write_header(ws, NON_ISSUE_COLUMNS)
    ws.freeze_panes = "A2"
    wrap = Alignment(wrap_text=True, vertical="top")
    for row_idx, finding in enumerate(_sorted(findings), start=2):
        values = [
            finding["id"],
            finding["module"],
            finding["file"],
            _location(finding),
            finding["title"],
            f"{finding['category']} {category_descriptions.get(finding['category'], '')}".strip(),
            finding["severity"],
            STATUS_ZH.get(_verifier_status(finding), _verifier_status(finding)),
            _verifier_reason(finding),
            _verifier_evidence_check(finding),
            finding["description"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap
    ws.auto_filter.ref = ws.dimensions


def _write_header(ws: Any, columns: list[tuple[str, int]]) -> None:
    _write_table_header(ws, 1, columns)


def _write_table_header(ws: Any, row: int, columns: list[tuple[str, int]]) -> None:
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    for col_idx, (name, width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _column_index(columns: list[tuple[str, int]], name: str) -> int:
    for i, (col_name, _) in enumerate(columns, start=1):
        if col_name == name:
            return i
    raise KeyError(name)


def _status_counter(findings: list[dict[str, Any]]) -> Counter[str]:
    return Counter(_verifier_status(finding) for finding in findings)


def _cli(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="render_xlsx",
        description="Render code-audit findings into a Chinese Excel workbook.",
    )
    parser.add_argument(
        "--workspace",
        type=Path,
        default=Path.cwd(),
        help="Workspace root (default: cwd). Used with --run-id.",
    )
    parser.add_argument(
        "--run-id",
        type=str,
        default=None,
        help="Run ID under .garage/code-audit/runs/<run-id>/.",
    )
    parser.add_argument(
        "--mode",
        choices=REPORT_MODES,
        default="auto",
        help="Report lifecycle mode: draft after review, final after verification, or auto.",
    )
    parser.add_argument(
        "--confirmed-path",
        type=Path,
        default=None,
        help="Explicit path to confirmed.json, used for final mode consistency checks.",
    )
    parser.add_argument("--plan-path", type=Path, default=None, help="Explicit path to plan.json.")
    parser.add_argument(
        "--findings-dir",
        type=Path,
        default=None,
        help="Explicit path to findings/ directory.",
    )
    parser.add_argument("--output", type=Path, default=None, help="Explicit output xlsx path.")
    parser.add_argument(
        "--allow-missing-openpyxl",
        action="store_true",
        help="Return success with a skipped result if openpyxl is unavailable.",
    )
    args = parser.parse_args(argv)

    try:
        result = render_workbook(
            workspace_root=args.workspace,
            run_id=args.run_id,
            confirmed_path=args.confirmed_path,
            plan_path=args.plan_path,
            findings_dir=args.findings_dir,
            output_path=args.output,
            mode=args.mode,
            strict=not args.allow_missing_openpyxl,
        )
    except ReportError as exc:
        print(f"render_xlsx: error: {exc}", file=sys.stderr)
        return 2

    if result.skipped:
        print(f"render_xlsx: skipped ({result.skipped_reason})", file=sys.stderr)
        return 0

    print(
        f"Wrote {result.output_path} "
        f"(mode={result.mode}, findings={result.finding_count}, "
        f"confirmed={result.confirmed_count}, non_issues={result.rejected_count}, "
        f"needs_more_evidence={result.needs_more_evidence_count})",
        file=sys.stderr,
    )
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(_cli())
