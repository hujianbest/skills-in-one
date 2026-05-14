#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
excel_helper.py — Pass 4 中文 Excel 报告生成器 (mdc-bug-patterns)

输入:
  --bugs-file findings.json     (符合 references/reporting.md 的 finding 列表)
  --coverage  coverage.json     可选, coverage_tracker.py 的 db, 用于覆盖率/总览
  --repo, --scope, --reviewer   可选元数据, 写入「审查总览」sheet
  --output    report.xlsx       默认 bug_report.xlsx

输出 (单一 .xlsx, 四个 sheet, 全中文表头, 适合人工逐条复核):

  1. 审查总览        基本信息 + 整体统计 + 阅读指引
  2. 发现明细        高/中可信发现, 一行一条, 带「人工确认」下拉
  3. 审计盲区        低可信 / 不确定 / 已登记但未审的候选
  4. 覆盖率明细      按模板与按文件的候选/确认/抑制/不确定统计

用法:
  excel_helper.py --bugs-file findings.json --coverage coverage.json \\
                  --repo my/repo --scope src/ --reviewer alice \\
                  --output bug_report.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:  # pragma: no cover
    sys.stderr.write("openpyxl 缺失, 请先 `pip install openpyxl`\n")
    raise


SEVERITY_ZH = {"critical": "严重", "high": "高", "medium": "中", "low": "低"}
CONFIDENCE_ZH = {"high": "高", "medium": "中", "low": "低"}
CATEGORY_ZH = {
    "memory": "内存安全",
    "null": "空指针/可空类型",
    "resource": "资源管理",
    "concurrency": "并发",
    "logic": "逻辑/数值",
}
VERDICT_ZH = {
    "agree":     "同意",
    "disagree":  "反对 (误报)",
    "uncertain": "不确定",
    "missing":   "未复核",
    "":          "未复核",
}
VERDICT_FILL = {
    "agree":     "FFB6E2B6",  # 浅绿
    "disagree":  "FFF2A6A6",  # 浅红
    "uncertain": "FFF7E18C",  # 浅黄
    "missing":   "FFD9D9D9",  # 灰
    "":          "FFD9D9D9",
}
SEVERITY_ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3}
CONFIDENCE_ORDER = {"high": 0, "medium": 1, "low": 2}

# 行底色 (淡, 不刺眼; 主要靠「严重程度」单元格本身的浓色块)
ROW_FILL = {
    "critical": "FFFFE6E6",
    "high":     "FFFFF1DC",
    "medium":   "FFFFFBE0",
    "low":      "FFEAF7E0",
}
# 严重程度单元格的浓色块
SEVERITY_FILL = {
    "critical": "FFE05656",
    "high":     "FFEE9933",
    "medium":   "FFE5C100",
    "low":      "FF6FAE3E",
}
SEVERITY_FONT_COLOR = {
    "critical": "FFFFFFFF",
    "high":     "FFFFFFFF",
    "medium":   "FF000000",
    "low":      "FFFFFFFF",
}

HEADER_FILL = PatternFill(start_color="FF305496", end_color="FF305496",
                          fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(border_style="thin", color="FFBFBFBF")
CELL_BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
WRAP_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
MONO = Font(name="Consolas", size=10)
META_LABEL = Font(bold=True, size=11)


# 发现明细 / 审计盲区 共用的列定义
#
# 列序原则: 让评审者从左到右先回答 "什么 → 多严重/多可信 → 在哪 → 证据 → 修复 → 复核",
# 因此 「问题说明」 紧随 编号/严重程度/可信度, 是首个内容列。
FINDING_COLUMNS = [
    ("编号",                   5),
    ("严重程度",               8),
    ("可信度",                 7),
    ("问题说明 (具体问题是什么)", 56),  # 首个内容列, 一句话说明这条 finding 究竟在说什么
    ("类别",                   12),
    ("模板ID",                 30),
    ("文件",                   36),   # 干净路径, 便于 git blame / CODEOWNERS 查责任人
    ("行号",                    8),
    ("所在函数",               22),
    ("证据 (file:line + 代码)", 60),
    ("已排除的误报模式",       28),
    ("修复建议",               46),
    ("代码上下文 (>>为问题行)", 60),
    ("子代理复核结论",         16),
    ("子代理复核依据",         60),
    ("人工确认",               22),
    ("备注",                   22),
]
# 列号常量 (1-based)
COL_SEV       = 2
COL_SUMMARY   = 4
COL_FILE      = 7
COL_LINE      = 8
COL_VERDICT   = 14
COL_VERDICT_R = 15
COL_CONFIRM   = 16
COL_NOTE      = 17


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--bugs-file", required=True,
                    help="findings JSON 列表 (见 references/reporting.md)")
    ap.add_argument("--coverage",
                    help="可选, coverage_tracker.py 的 coverage.json")
    ap.add_argument("--output", "-o", default="bug_report.xlsx",
                    help="输出 .xlsx 路径")
    ap.add_argument("--repo", default="",
                    help="仓库标识 (会写入「审查总览」)")
    ap.add_argument("--scope", default="",
                    help="审计范围 (会写入「审查总览」)")
    ap.add_argument("--reviewer", default="",
                    help="审计人 (会写入「审查总览」)")
    args = ap.parse_args(argv)

    bugs = json.loads(Path(args.bugs_file).read_text(encoding="utf-8"))
    if not isinstance(bugs, list):
        sys.stderr.write("findings 文件必须是 JSON 数组\n")
        return 2

    coverage = None
    if args.coverage:
        cov_path = Path(args.coverage)
        if cov_path.exists():
            coverage = json.loads(cov_path.read_text(encoding="utf-8"))
        else:
            sys.stderr.write(f"warn: coverage 文件不存在, 略过: {cov_path}\n")

    findings, gaps = _split_findings(bugs)

    wb = Workbook()
    ws_overview = wb.active
    ws_overview.title = "审查总览"
    ws_findings = wb.create_sheet("发现明细")
    ws_gaps = wb.create_sheet("审计盲区")
    ws_coverage = wb.create_sheet("覆盖率明细")

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    _write_overview(ws_overview, args, findings, gaps, coverage, timestamp)
    _write_findings_sheet(ws_findings, findings,
                          empty_msg="本表为空: 当前没有高/中可信的发现。")
    _write_findings_sheet(ws_gaps, gaps,
                          empty_msg="本表为空: 没有需要进一步确认的低可信或不确定项。",
                          gap_mode=True)
    _write_coverage_sheet(ws_coverage, bugs, coverage)

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"已生成中文审查报告: {args.output}")
    print(f"  发现明细: {len(findings)} 条 (高/中可信)")
    print(f"  审计盲区: {len(gaps)} 条 (低/不确定)")
    return 0


# ---------- helpers ----------------------------------------------------------


def _split_findings(bugs: list[dict[str, Any]]) -> tuple[list[dict], list[dict]]:
    findings, gaps = [], []
    for b in bugs:
        conf = str(b.get("confidence", "low")).lower()
        status = str(b.get("status", "confirmed")).lower()
        if conf == "low" or status in ("inconclusive", "open"):
            gaps.append(b)
        else:
            findings.append(b)
    findings.sort(key=_sort_key)
    gaps.sort(key=_sort_key)
    return findings, gaps


def _sort_key(b: dict[str, Any]) -> tuple:
    loc = b.get("location") or {}
    return (
        SEVERITY_ORDER.get(str(b.get("severity", "low")).lower(), 9),
        CONFIDENCE_ORDER.get(str(b.get("confidence", "low")).lower(), 9),
        str(loc.get("file", "")),
        int(loc.get("line", 0) or 0),
    )


def _location(b: dict[str, Any]) -> tuple[str, Any, str]:
    """Return (file, line, function). `line` may be int or '' for sortable cells."""
    loc = b.get("location") or {}
    if isinstance(loc, dict):
        line = loc.get("line", "")
        try:
            line = int(line)
        except (ValueError, TypeError):
            line = line or ""
        return (
            str(loc.get("file", "") or ""),
            line,
            str(loc.get("function", "") or ""),
        )
    return (str(loc), "", "")


def _verdict_summary(b: dict[str, Any]) -> tuple[str, str]:
    """Return (verdict_label_zh, rationale_text)."""
    rev = b.get("second_pass_review") or {}
    verdict = str(rev.get("verdict", "") or "").lower()
    label = VERDICT_ZH.get(verdict, "未复核")
    parts: list[str] = []
    if rev.get("rationale"):
        parts.append(str(rev["rationale"]))
    ec = rev.get("evidence_check") or {}
    if ec:
        check_lines: list[str] = []
        check_lines.append("evidence_check:")
        for k in (
            "all_cited_lines_exist",
            "all_cited_lines_match_excerpts",
            "fp_filters_actually_ruled_out",
        ):
            v = ec.get(k)
            if v is True:
                mark = "✓"
            elif v is False:
                mark = "✗"
            else:
                mark = "?"
            check_lines.append(f"  {mark} {k}")
        extra = ec.get("additional_fp_filters_found") or []
        if extra:
            check_lines.append(
                "  + additional_fp_filters_found: " + ", ".join(extra)
            )
        parts.append("\n".join(check_lines))
    sup = rev.get("supporting_evidence") or []
    if sup:
        parts.append("supporting_evidence:\n" +
                     "\n".join(f"  • {x}" for x in sup))
    if rev.get("reviewer") or rev.get("reviewed_at"):
        parts.append(f"by {rev.get('reviewer', '?')} @ {rev.get('reviewed_at', '?')}")
    if not parts:
        parts.append("(暂无子代理复核结论)")
    return label, "\n\n".join(parts)


def _join_evidence(b: dict[str, Any]) -> str:
    ev = b.get("required_evidence") or b.get("evidence") or {}
    if isinstance(ev, dict):
        return "\n".join(f"• {k}: {v}" for k, v in ev.items())
    if isinstance(ev, list):
        return "\n".join(f"• {x}" for x in ev)
    return str(ev or "")


def _join_list(field: Any, bullet: str = "• ") -> str:
    if isinstance(field, list):
        return "\n".join(f"{bullet}{x}" for x in field) if field else ""
    return str(field or "")


def _join_context(b: dict[str, Any]) -> str:
    ctx = b.get("context") or []
    if isinstance(ctx, list):
        return "\n".join(str(x) for x in ctx)
    return str(ctx or "")


# ---------- sheets -----------------------------------------------------------


def _write_overview(
    ws,
    args: argparse.Namespace,
    findings: list[dict],
    gaps: list[dict],
    coverage: dict | None,
    timestamp: str,
) -> None:
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12

    title = ws.cell(1, 1, "C/C++ 代码 Bug 审查报告")
    title.font = Font(bold=True, size=18, color="FF1F3864")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.row_dimensions[1].height = 30

    row = 3
    ws.cell(row, 1, "基本信息").font = Font(bold=True, size=13,
                                          color="FF1F3864")
    row += 1
    for label, value in [
        ("仓库 / 分支", args.repo or "(未指定)"),
        ("审计范围", args.scope or "(未指定)"),
        ("审计人", args.reviewer or "(未指定)"),
        ("审计时间", timestamp),
        ("审计模板数",
         str(len({b.get("template_id") for b in findings + gaps
                 if b.get("template_id")}))),
    ]:
        ws.cell(row, 1, label).font = META_LABEL
        ws.cell(row, 2, value)
        row += 1

    if coverage and isinstance(coverage.get("candidates"), dict):
        ws.cell(row, 1, "已审候选数 / 总候选数").font = META_LABEL
        all_cs = coverage["candidates"].values()
        n_total = len(coverage["candidates"])
        n_done = sum(1 for c in all_cs
                     if c.get("status") in ("confirmed", "suppressed"))
        ws.cell(row, 2, f"{n_done} / {n_total}")
        row += 1

    row += 1
    ws.cell(row, 1, "发现统计").font = Font(bold=True, size=13,
                                          color="FF1F3864")
    row += 1
    headers = ["严重程度", "高可信", "中可信", "审计盲区(低/待定)", "合计"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row, col, h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
        c.border = CELL_BORDER
    row += 1

    by_sev = defaultdict(lambda: defaultdict(int))
    by_verdict = defaultdict(int)
    for b in findings:
        by_sev[str(b.get("severity", "")).lower()][str(b.get("confidence", "")).lower()] += 1
        v = str((b.get("second_pass_review") or {}).get("verdict", "") or "").lower()
        by_verdict[v or "missing"] += 1
    for b in gaps:
        by_sev[str(b.get("severity", "")).lower()]["gap"] += 1
        v = str((b.get("second_pass_review") or {}).get("verdict", "") or "").lower()
        by_verdict[v or "missing"] += 1

    totals = defaultdict(int)
    for sev in ("critical", "high", "medium", "low"):
        h = by_sev[sev].get("high", 0)
        m = by_sev[sev].get("medium", 0)
        g = by_sev[sev].get("low", 0) + by_sev[sev].get("gap", 0)
        total = h + m + g
        sev_cell = ws.cell(row, 1, SEVERITY_ZH.get(sev, sev))
        sev_cell.fill = PatternFill(
            start_color=SEVERITY_FILL.get(sev, "FFCCCCCC"),
            end_color=SEVERITY_FILL.get(sev, "FFCCCCCC"),
            fill_type="solid")
        sev_cell.font = Font(
            bold=True,
            color=SEVERITY_FONT_COLOR.get(sev, "FF000000"))
        sev_cell.alignment = CENTER
        ws.cell(row, 2, h).alignment = CENTER
        ws.cell(row, 3, m).alignment = CENTER
        ws.cell(row, 4, g).alignment = CENTER
        ws.cell(row, 5, total).alignment = CENTER
        for col in range(1, 6):
            ws.cell(row, col).border = CELL_BORDER
        totals["high"] += h
        totals["medium"] += m
        totals["gap"] += g
        totals["all"] += total
        row += 1
    # 合计行
    ws.cell(row, 1, "合计").font = Font(bold=True)
    ws.cell(row, 1).alignment = CENTER
    ws.cell(row, 2, totals["high"]).alignment = CENTER
    ws.cell(row, 3, totals["medium"]).alignment = CENTER
    ws.cell(row, 4, totals["gap"]).alignment = CENTER
    ws.cell(row, 5, totals["all"]).alignment = CENTER
    for col in range(1, 6):
        ws.cell(row, col).border = CELL_BORDER
        ws.cell(row, col).fill = PatternFill(
            start_color="FFE7E6E6", end_color="FFE7E6E6", fill_type="solid")
    row += 2

    # 子代理复核统计
    ws.cell(row, 1, "子代理复核统计").font = Font(bold=True, size=13,
                                                color="FF1F3864")
    row += 1
    verdict_headers = [("同意", "agree"), ("反对 (误报)", "disagree"),
                       ("不确定", "uncertain"), ("未复核", "missing")]
    for col, (label, _) in enumerate(verdict_headers, 1):
        c = ws.cell(row, col, label)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
        c.border = CELL_BORDER
    ws.cell(row, 5, "合计").fill = HEADER_FILL
    ws.cell(row, 5).font = HEADER_FONT
    ws.cell(row, 5).alignment = HEADER_ALIGN
    ws.cell(row, 5).border = CELL_BORDER
    row += 1
    v_total = 0
    for col, (label, key) in enumerate(verdict_headers, 1):
        n = by_verdict.get(key, 0)
        v_total += n
        cell = ws.cell(row, col, n)
        cell.alignment = CENTER
        cell.border = CELL_BORDER
        fill_color = VERDICT_FILL.get(key, VERDICT_FILL[""])
        cell.fill = PatternFill(start_color=fill_color,
                                end_color=fill_color, fill_type="solid")
    total_cell = ws.cell(row, 5, v_total)
    total_cell.alignment = CENTER
    total_cell.border = CELL_BORDER
    total_cell.fill = PatternFill(
        start_color="FFE7E6E6", end_color="FFE7E6E6", fill_type="solid")
    total_cell.font = Font(bold=True)
    row += 2

    ws.cell(row, 1, "阅读指引").font = Font(bold=True, size=13,
                                          color="FF1F3864")
    row += 1
    for line in [
        "1. 打开「发现明细」页后, 请按 编号 → 严重程度 → 可信度 → 「问题说明 (具体问题是什么)」 的顺序读每一行, 一句话先弄清「这条 finding 在说什么」。",
        "2. 然后阅读「证据」「代码上下文」核对; 用「文件」「行号」两列直接 git blame 或在 CODEOWNERS 中查责任人。",
        "3. 「子代理复核结论」是另一只 AI 子代理独立复核的判断 (绿=同意 / 红=反对 / 黄=不确定 / 灰=未复核); 与原结论分歧时请重点核对「子代理复核依据」列。",
        "4. 在最后两列填写「人工确认」(下拉: ✓ 同意 / ✗ 误报 / ? 待定) 与「备注」。",
        "5. 「审计盲区」页含低可信与不确定项 — 这些不是确认的 bug, 但本次审计未能完全排除, 请按需追加审查。",
        "6. 「覆盖率明细」页给出每个模板与每个文件的候选/确认/抑制/不确定计数, 用于评估本次审计的充分度。",
        "7. 严重程度仅表示「若属实, 后果有多严重」; 可信度表示「我们有多确定它属实」。两者独立, 都需关注。",
        "8. 「已排除的误报模式」列出审计时主动验证并排除的 FP 过滤器 (fp.* 命名), 便于复核者验证排除是否成立。",
    ]:
        cell = ws.cell(row, 1, line)
        cell.alignment = WRAP_TOP
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 26
        row += 1


def _write_findings_sheet(
    ws,
    rows: list[dict[str, Any]],
    *,
    empty_msg: str = "",
    gap_mode: bool = False,
) -> None:
    # header
    for col, (h, w) in enumerate(FINDING_COLUMNS, 1):
        c = ws.cell(1, col, h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
        c.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "C2"  # 冻结表头 + 编号 + 严重程度
    ws.auto_filter.ref = f"A1:{get_column_letter(len(FINDING_COLUMNS))}1"

    if not rows:
        c = ws.cell(2, 1, empty_msg)
        c.font = Font(italic=True, color="FF7F7F7F")
        ws.merge_cells(start_row=2, start_column=1,
                       end_row=2, end_column=len(FINDING_COLUMNS))
        return

    # 人工确认列 数据校验下拉
    confirm_letter = get_column_letter(COL_CONFIRM)
    dv = DataValidation(
        type="list",
        formula1='"✓ 同意 (确认是bug),✗ 误报 (附理由),? 待定 (需更多上下文)"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "请从下拉列表中选择一项"
    dv.errorTitle = "无效的人工确认值"
    dv.prompt = "请逐条审核后填写"
    dv.promptTitle = "人工确认"
    ws.add_data_validation(dv)
    dv.add(f"{confirm_letter}2:{confirm_letter}{1 + len(rows)}")

    for i, b in enumerate(rows, start=1):
        row_idx = i + 1
        sev = str(b.get("severity", "")).lower()
        conf = str(b.get("confidence", "")).lower()
        cat = str(b.get("category", "")).lower()
        file_text, line_val, func_text = _location(b)
        verdict_label, verdict_rationale = _verdict_summary(b)
        verdict_key = str(
            (b.get("second_pass_review") or {}).get("verdict", "") or ""
        ).lower()

        summary_text = b.get("summary") or b.get("name") or "(未提供问题说明)"

        cells = [
            ("",        str(i)),                                    # 1  编号
            ("",        SEVERITY_ZH.get(sev, sev or "")),           # 2  严重程度
            ("",        CONFIDENCE_ZH.get(conf, conf or "")),       # 3  可信度
            ("summary", summary_text),                              # 4  问题说明
            ("",        CATEGORY_ZH.get(cat, cat or "")),           # 5  类别
            ("",        b.get("template_id", "")),                  # 6  模板ID
            ("",        file_text),                                 # 7  文件
            ("",        line_val),                                  # 8  行号
            ("",        func_text),                                 # 9  所在函数
            ("ev",      _join_evidence(b)),                         # 10 证据
            ("",        _join_list(b.get("false_positive_filters_ruled_out"))),  # 11
            ("",        _join_list(b.get("fix_suggestions"))),      # 12 修复建议
            ("ev",      _join_context(b)),                          # 13 代码上下文
            ("",        verdict_label),                             # 14 子代理复核结论
            ("ev",      verdict_rationale),                         # 15 子代理复核依据
            ("",        ""),                                        # 16 人工确认
            ("",        ""),                                        # 17 备注
        ]
        center_cols = {1, COL_SEV, 3, COL_LINE, COL_VERDICT}
        for col_idx, (kind, value) in enumerate(cells, start=1):
            c = ws.cell(row_idx, col_idx, value)
            c.border = CELL_BORDER
            c.alignment = CENTER if col_idx in center_cols else WRAP_TOP
            if kind == "ev":
                c.font = MONO
            elif kind == "summary":
                # 问题说明 用稍大、加粗一点的字体, 让人一眼看到本条 finding 在说什么
                c.font = Font(size=11, bold=True, color="FF1F3864")

        # 严重程度 单独着色 (浓色块)
        sev_cell = ws.cell(row_idx, COL_SEV)
        sev_cell.fill = PatternFill(
            start_color=SEVERITY_FILL.get(sev, "FFCCCCCC"),
            end_color=SEVERITY_FILL.get(sev, "FFCCCCCC"),
            fill_type="solid")
        sev_cell.font = Font(
            bold=True,
            color=SEVERITY_FONT_COLOR.get(sev, "FF000000"))

        # 整行淡底色
        light = ROW_FILL.get(sev)
        if light:
            light_fill = PatternFill(start_color=light, end_color=light,
                                     fill_type="solid")
            for col_idx in range(1, len(FINDING_COLUMNS) + 1):
                if col_idx in (COL_SEV, COL_VERDICT,
                               COL_CONFIRM, COL_NOTE):
                    continue  # 这些列单独着色或保留留白
                ws.cell(row_idx, col_idx).fill = light_fill

        # 子代理复核结论 单独着色 (绿/红/黄/灰)
        verdict_color = VERDICT_FILL.get(verdict_key, VERDICT_FILL[""])
        v_cell = ws.cell(row_idx, COL_VERDICT)
        v_cell.fill = PatternFill(
            start_color=verdict_color, end_color=verdict_color,
            fill_type="solid",
        )
        v_cell.font = Font(bold=True)

        # 行高: 依据证据/上下文/复核依据/问题说明行数估算
        # 问题说明假设按 col 4 宽度 (~56 字符) 自动换行
        summary_wrapped = max(1, (len(summary_text) // 50) + summary_text.count("\n") + 1)
        ev_lines = max(
            _line_count(_join_evidence(b)),
            _line_count(_join_context(b)),
            _line_count(_join_list(b.get("fix_suggestions"))),
            _line_count(verdict_rationale),
            summary_wrapped,
        )
        ws.row_dimensions[row_idx].height = max(60, min(420, 16 * (ev_lines + 1)))

        if gap_mode:
            # 给「备注」一个友好的提示 (不写入值; 仅文本)
            note = b.get("reason") or b.get("note") or ""
            if note:
                ws.cell(row_idx, COL_NOTE, f"原因: {note}")


def _write_coverage_sheet(ws, bugs: list[dict], coverage: dict | None) -> None:
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    for col in "DEFGH":
        ws.column_dimensions[col].width = 12

    row = 1
    title = ws.cell(row, 1, "按模板的覆盖统计")
    title.font = Font(bold=True, size=13, color="FF1F3864")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

    headers = ["模板ID", "类别", "默认严重",
               "候选数", "已确认", "已抑制", "不确定", "覆盖率"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row, col, h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
        c.border = CELL_BORDER
    row += 1

    # 优先用 coverage_tracker 的真实数据; 否则退化到 findings 里的统计
    by_template: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"candidates": 0, "confirmed": 0, "suppressed": 0,
                 "inconclusive": 0, "open": 0,
                 "category": "", "severity": ""}
    )

    if coverage and isinstance(coverage.get("candidates"), dict):
        for c in coverage["candidates"].values():
            tid = c.get("template_id") or "<未知>"
            t = by_template[tid]
            t["candidates"] += 1
            status = c.get("status", "open")
            if status in t:
                t[status] += 1
            else:
                t["open"] += 1
    # 类别/严重 从 bugs 推断
    for b in bugs:
        tid = b.get("template_id")
        if not tid:
            continue
        t = by_template[tid]
        if not t["category"]:
            t["category"] = b.get("category", "")
        if not t["severity"]:
            t["severity"] = b.get("severity", "")
        if not coverage:
            t["candidates"] += 1
            t[("confirmed" if str(b.get("confidence", "low")).lower() != "low"
               else "inconclusive")] += 1

    if not by_template:
        ws.cell(row, 1, "(无数据)").font = Font(italic=True,
                                              color="FF7F7F7F")
        return

    for tid in sorted(by_template):
        t = by_template[tid]
        n = t["candidates"]
        cov = ((t["confirmed"] + t["suppressed"]) / n * 100.0) if n else 0.0
        ws.cell(row, 1, tid)
        ws.cell(row, 2, CATEGORY_ZH.get(t["category"], t["category"] or ""))
        ws.cell(row, 3, SEVERITY_ZH.get(t["severity"], t["severity"] or ""))
        ws.cell(row, 4, n)
        ws.cell(row, 5, t["confirmed"])
        ws.cell(row, 6, t["suppressed"])
        ws.cell(row, 7, t["inconclusive"] + t["open"])
        ws.cell(row, 8, f"{cov:.1f}%")
        for col in range(1, 9):
            ws.cell(row, col).border = CELL_BORDER
            ws.cell(row, col).alignment = (
                WRAP_TOP if col in (1, 2) else CENTER
            )
        row += 1
    row += 2

    # ---- 按文件 ----
    ws.cell(row, 1, "按文件的发现统计").font = Font(bold=True, size=13,
                                                color="FF1F3864")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1
    headers2 = ["文件", "高可信发现", "中可信发现", "低/不确定", "总发现"]
    for col, h in enumerate(headers2, 1):
        c = ws.cell(row, col, h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
        c.border = CELL_BORDER
    row += 1

    by_file: dict[str, dict[str, int]] = defaultdict(
        lambda: {"high": 0, "medium": 0, "low_or_inc": 0})
    for b in bugs:
        loc = b.get("location") or {}
        f = (loc.get("file") if isinstance(loc, dict) else "") or "(unknown)"
        conf = str(b.get("confidence", "low")).lower()
        status = str(b.get("status", "")).lower()
        if conf == "low" or status in ("inconclusive", "open"):
            by_file[f]["low_or_inc"] += 1
        elif conf == "medium":
            by_file[f]["medium"] += 1
        else:
            by_file[f]["high"] += 1

    if not by_file:
        ws.cell(row, 1, "(无数据)").font = Font(italic=True,
                                              color="FF7F7F7F")
        return

    for f in sorted(by_file):
        v = by_file[f]
        total = v["high"] + v["medium"] + v["low_or_inc"]
        ws.cell(row, 1, f)
        ws.cell(row, 2, v["high"])
        ws.cell(row, 3, v["medium"])
        ws.cell(row, 4, v["low_or_inc"])
        ws.cell(row, 5, total)
        for col in range(1, 6):
            ws.cell(row, col).border = CELL_BORDER
            ws.cell(row, col).alignment = (
                WRAP_TOP if col == 1 else CENTER
            )
        row += 1


def _line_count(s: str) -> int:
    if not s:
        return 1
    return s.count("\n") + 1


if __name__ == "__main__":
    sys.exit(main())
